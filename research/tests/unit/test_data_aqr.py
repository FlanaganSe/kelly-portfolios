"""Offline parser tests for the AQR workbook reader.

``tests/fixtures/aqr_tsmom_sample.xlsx`` is a small workbook whose **content** is
copied verbatim out of a real download of 2026-08-12: the same prose preamble in
the same rows, the same header row in the same position, the same month-end
datetimes and the same return values, plus the same four sheets with the same
almost-empty prose tabs. Its *bytes* are not the vendor's bytes — an ``.xlsx``
cannot be sliced the way a CSV can, so the fixture was rewritten by openpyxl. The
live download is covered by the ``network`` test in
``tests/integration/test_data_network.py``.

``tests/fixtures/aqr_drawing_sample.emf`` is the real ``xl/media/image2.emf``
extracted verbatim from the same download, so the drawing-text recovery is tested
against genuine vendor bytes rather than a synthesised record stream.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
import pytest

from portfolio_edge.data import aqr
from portfolio_edge.data.cache import CacheEntry, RawArtifactMissing, RawCache
from portfolio_edge.data.validation import validate_table

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
WORKBOOK = FIXTURES / "aqr_tsmom_sample.xlsx"
DRAWING = FIXTURES / "aqr_drawing_sample.emf"


@pytest.fixture
def dataset() -> aqr.AqrDataset:
    return aqr.get_dataset("aqr_tsmom_factors")


def _seed(cache: RawCache, payload: bytes, url: str) -> CacheEntry:
    return cache.store(
        url,
        payload,
        headers={
            "Content-Type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            "Last-Modified": "Fri, 26 Jun 2026 15:54:00 GMT",
            "Content-Disposition": 'attachment; filename="Time Series Momentum.xlsx"',
        },
        retrieved_utc="2026-08-12T00:00:00Z",
    )


@pytest.fixture
def parsed(tmp_path: Path, dataset: aqr.AqrDataset) -> tuple[RawCache, CacheEntry, aqr.AqrFile]:
    cache = RawCache(tmp_path)
    entry = _seed(cache, WORKBOOK.read_bytes(), dataset.url)
    return (cache, entry, aqr.parse(cache, entry, dataset=dataset))


def test_parsing_requires_a_cached_raw_artifact(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    cache = RawCache(tmp_path)
    entry = CacheEntry(
        url=dataset.url,
        sha256="0" * 64,
        size_bytes=0,
        retrieved_utc="2026-08-12T00:00:00Z",
        http_status=200,
        headers=(),
    )
    with pytest.raises(RawArtifactMissing):
        aqr.parse(cache, entry, dataset=dataset)


def test_the_declared_sheet_is_read_and_every_sheet_is_recorded(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    _, _, file = parsed
    assert file.data_sheet == "TSMOM Factors"
    assert file.sheet_names == (
        "TSMOM Factors",
        "Definitions",
        "Data Sources",
        "Disclosures",
    )


def test_a_renamed_sheet_is_refused_rather_than_guessed_around(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    """Falling back to another sheet would change every number invisibly."""
    cache = RawCache(tmp_path)
    entry = _seed(cache, WORKBOOK.read_bytes(), dataset.url)
    renamed = aqr.AqrDataset(
        dataset_id=dataset.dataset_id,
        filename=dataset.filename,
        data_sheet="TSMOM Factors (2026 revision)",
        description=dataset.description,
        declared_source_units=dataset.declared_source_units,
        declared_units=dataset.declared_units,
        declared_unit_transform=dataset.declared_unit_transform,
        declared_return_basis=dataset.declared_return_basis,
        availability_policy=dataset.availability_policy,
        revision_policy=dataset.revision_policy,
    )
    with pytest.raises(aqr.AqrSheetMissingError, match="TSMOM Factors"):
        aqr.parse(cache, entry, dataset=renamed)


def test_the_data_block_is_found_structurally_not_by_row_offset(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    """AQR edits the disclaimer block, so a hardcoded offset is wrong by construction."""

    workbook = openpyxl.load_workbook(WORKBOOK)
    sheet = workbook["TSMOM Factors"]
    sheet.insert_rows(1, amount=7)
    for offset in range(7):
        sheet.cell(row=offset + 1, column=1, value=f"An extra disclaimer line {offset}.")
    payload = io.BytesIO()
    workbook.save(payload)

    cache = RawCache(tmp_path)
    entry = _seed(cache, payload.getvalue(), dataset.url + "#padded")
    shifted = aqr.parse(cache, entry, dataset=dataset)
    original = aqr.parse(
        cache, _seed(cache, WORKBOOK.read_bytes(), dataset.url), dataset=dataset
    )
    assert shifted.table.periods == original.table.periods
    assert shifted.table.values == original.table.values
    assert "An extra disclaimer line 0." in shifted.preamble


def test_month_end_dates_become_period_labels_and_the_frequency_is_measured(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    _, _, file = parsed
    assert file.table.frequency == "monthly"
    assert file.table.table_id == "monthly"
    assert file.table.first_observation == "1985-01"
    assert file.table.last_observation == "1986-02"
    assert file.table.rows == 14


def test_columns_come_from_the_header_row_and_trailing_blanks_are_dropped(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    """The sheet's declared dimension runs wider than the data it holds."""
    _, _, file = parsed
    assert file.table.columns == ("TSMOM", "TSMOM^CM", "TSMOM^EQ", "TSMOM^FI", "TSMOM^FX")


def test_values_are_not_transformed_and_the_declaration_is_cross_checked(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    """The workbook declares no units, so the parser states a claim and checks it."""
    _, _, file = parsed
    assert file.table.source_units == "decimal"
    assert file.table.unit_transform == "identity"
    # 1985-01 TSMOM is 0.043456226781221075 in the real workbook.
    assert file.table.values[0][0] == pytest.approx(0.043456226781221075)
    joined = " ".join(file.table.warnings)
    assert "DECLARED by this repository" in joined
    assert "declares no units anywhere in machine-readable form" in joined


def test_an_excess_return_series_is_flagged_as_one(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    _, _, file = parsed
    assert any("EXCESS returns" in warning for warning in file.table.warnings)


def test_a_percent_magnitude_in_a_decimal_declaration_is_a_loud_contradiction(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    """A file read at the wrong scale survives every other check."""

    workbook = openpyxl.load_workbook(WORKBOOK)
    sheet = workbook["TSMOM Factors"]
    for row in sheet.iter_rows(min_row=19, max_row=32, min_col=2, max_col=6):
        for cell in row:
            if isinstance(cell.value, float):
                cell.value = cell.value * 100.0
    payload = io.BytesIO()
    workbook.save(payload)

    cache = RawCache(tmp_path)
    entry = _seed(cache, payload.getvalue(), dataset.url + "#percent")
    file = aqr.parse(cache, entry, dataset=dataset)
    joined = " ".join(file.table.warnings)
    assert "CONTRADICTION" in joined
    # Nothing was divided by anything.
    assert file.table.values[0][0] == pytest.approx(4.3456226781221075)


def test_the_prose_only_sheets_are_reported_as_carrying_no_readable_text(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile],
) -> None:
    """The vendor's methodology lives in pictures; a cell reader sees nothing."""
    _, _, file = parsed
    joined = " ".join(file.warnings)
    assert "NOT machine-readable" in joined
    assert "Definitions" in joined and "Disclosures" in joined


def test_drawing_text_is_recovered_from_real_vendor_bytes() -> None:
    payload = io.BytesIO()
    with zipfile.ZipFile(payload, "w") as archive:
        archive.writestr("xl/media/image1.emf", DRAWING.read_bytes())
    recovered = aqr.recover_drawing_text(payload.getvalue())

    assert len(recovered) == 1
    part, text = recovered[0]
    assert part == "xl/media/image1.emf"
    assert "Moskowitz" in text
    assert "Pedersen" in text
    assert "Time" in text and "Series" in text and "Momentum" in text


def test_a_workbook_with_no_drawings_recovers_nothing_rather_than_failing() -> None:
    assert aqr.recover_drawing_text(WORKBOOK.read_bytes()) == ()


def test_bytes_that_are_not_a_workbook_are_refused(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    cache = RawCache(tmp_path)
    entry = cache.store(dataset.url + "#html", b"<html>Access denied</html>")
    with pytest.raises(aqr.AqrParseError, match="not a zip container"):
        aqr.parse(cache, entry, dataset=dataset)


def test_a_sheet_with_no_date_keyed_rows_is_refused(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TSMOM Factors"
    sheet.append(["Some prose"])
    sheet.append([None, "TSMOM"])
    sheet.append(["not a date", 0.01])
    payload = io.BytesIO()
    workbook.save(payload)

    cache = RawCache(tmp_path)
    entry = cache.store(dataset.url + "#nodates", payload.getvalue())
    with pytest.raises(aqr.AqrParseError, match="no rows whose first cell is a date"):
        aqr.parse(cache, entry, dataset=dataset)


def test_a_bare_number_is_not_accepted_as_a_date(
    tmp_path: Path, dataset: aqr.AqrDataset
) -> None:
    """An unformatted Excel serial is indistinguishable from a return."""

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TSMOM Factors"
    sheet.append(["Some prose"])
    sheet.append([None, "TSMOM"])
    for serial in range(31048, 31060):
        sheet.append([serial, 0.01])
    payload = io.BytesIO()
    workbook.save(payload)

    cache = RawCache(tmp_path)
    entry = cache.store(dataset.url + "#serial", payload.getvalue())
    with pytest.raises(aqr.AqrParseError, match="no rows whose first cell is a date"):
        aqr.parse(cache, entry, dataset=dataset)


def test_the_table_passes_validation(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile], dataset: aqr.AqrDataset
) -> None:
    _, _, file = parsed
    report = validate_table(
        file.table,
        dataset_id="aqr_tsmom_factors_monthly",
        expected_columns=dataset.expected_columns,
        expected_frequency="monthly",
    )
    assert report.ok, report.summary()


def test_the_manifest_pins_the_sheet_and_the_revision_policy(
    parsed: tuple[RawCache, CacheEntry, aqr.AqrFile], dataset: aqr.AqrDataset
) -> None:
    _, entry, file = parsed
    manifests = aqr.build_manifests(dataset, entry, file)

    assert len(manifests) == 1
    manifest = manifests[0]
    assert manifest.dataset_id == "aqr_tsmom_factors_monthly"
    assert manifest.sha256_raw == entry.sha256
    assert manifest.sha256_normalized == file.table.sha256_normalized()
    assert manifest.rows == 14
    assert manifest.parser_version == aqr.PARSER_VERSION
    assert manifest.source_last_modified == "Fri, 26 Jun 2026 15:54:00 GMT"

    pinned = [w for w in manifest.warnings if w.startswith("SHEET PINNED:")]
    assert len(pinned) == 1
    assert "'TSMOM Factors'" in pinned[0]

    assert "not point-in-time" in manifest.revision_policy.lower()
    assert "reconstructs the full history" in manifest.revision_policy
    assert any("No fee, transaction-cost" in w for w in manifest.warnings)


def test_unknown_dataset_id_lists_the_known_ones() -> None:
    with pytest.raises(KeyError, match="aqr_tsmom_factors"):
        aqr.get_dataset("aqr_value_factors")
