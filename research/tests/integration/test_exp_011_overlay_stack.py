"""End-to-end tests for Experiment 011, offline.

The four real files cannot be committed, so the whole path -- cache, zip-or-CSV
extraction, two spreadsheet readers, the ``ltr - Rfree`` and ``corpr - Rfree``
differences, the hard sha256 pins, the intersection that defines the window, the six
portfolios, the costs charged on notional, the matched-volatility comparison, the
haircut sweep, the admission test, the artifacts and the ledger -- runs against
synthetic files in exactly the vendors' layouts, seeded into the real cache under the
real URLs.

**Expected values are computed in this file, with plain NumPy, from the generated
numbers**, never by calling the code under test.
"""

from __future__ import annotations

import datetime as dt
import io
import itertools
import math
import os
from collections.abc import Iterator, Mapping, Sequence
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl  # type: ignore[import-untyped]
import pytest
import yaml

from portfolio_edge.data import aqr, french, goyal_welch
from portfolio_edge.data.cache import CACHE_ENV_VAR, RawCache
from portfolio_edge.experiments.exp_011_overlay_stack import (
    ENTRY_POINT,
    MONTHS_PER_YEAR,
    OverlayStackError,
    build_registry,
    default_specification_path,
    load_panel,
    run,
)
from portfolio_edge.experiments.ledger import Ledger, LedgerEvent, RunStatus
from portfolio_edge.experiments.registry import RunContext
from portfolio_edge.experiments.result import ExperimentResult, ResultStatus
from portfolio_edge.experiments.runner import run_experiment
from portfolio_edge.experiments.specification import (
    Specification,
    specification_from_mapping,
)

SPEC_WINDOW = ("1985-01", "2025-05")
LONG_START = "1926-07"

#: Coverage of each synthetic file, chosen so the intersection is decided by the trend
#: file's start and the commodity file's end, exactly as the real files decide it.
COVERAGE = {
    "french": (LONG_START, "2026-06"),
    "goyal_welch": (LONG_START, "2025-12"),
    "commodity": (LONG_START, "2025-05"),
    "trend": ("1985-01", "2026-05"),
}

FRENCH_COLUMNS = ("Mkt-RF", "SMB", "HML", "RF")
GOYAL_WELCH_COLUMNS = ("tbl", "Rfree", "infl", "ltr", "corpr")
COMMODITY_COLUMN = "Excess return of equal-weight commodities portfolio"


# --------------------------------------------------------------------------- #
# Synthetic sources, in the real layouts
# --------------------------------------------------------------------------- #


def as_mapping(value: object) -> Mapping[str, Any]:
    assert isinstance(value, Mapping)
    return value


def number(value: object) -> float:
    assert isinstance(value, int | float) and not isinstance(value, bool)
    return float(value)


def month_keys(start: str, end: str) -> list[tuple[int, int]]:
    first = int(start[:4]) * 12 + int(start[5:]) - 1
    last = int(end[:4]) * 12 + int(end[5:]) - 1
    return [(index // 12, index % 12 + 1) for index in range(first, last + 1)]


def label(key: tuple[int, int]) -> str:
    return f"{key[0]:04d}-{key[1]:02d}"


def month_end(key: tuple[int, int]) -> dt.date:
    year, month = key
    following = dt.date(year + (month == 12), month % 12 + 1, 1)
    return following - dt.timedelta(days=1)


def french_csv(rows: Mapping[tuple[int, int], Sequence[float]]) -> bytes:
    """One Ken French factor file, percent, with the same preamble shape as the real one."""
    lines = [
        "This file was created using a synthetic fixture, not a CRSP database.",
        "Missing data are indicated by -99.99.",
        "",
        "," + ",".join(FRENCH_COLUMNS),
    ]
    for key, values in rows.items():
        cells = ",".join(f"{value * 100.0:>8.2f}" for value in values)
        lines.append(f"{key[0]:04d}{key[1]:02d},{cells}")
    lines.append("Copyright 2026 synthetic fixture")
    return ("\n".join(lines) + "\n").encode("utf-8")


def goyal_welch_xlsx(rows: Mapping[tuple[int, int], Sequence[float]]) -> bytes:
    """A predictor workbook with the three sheets the reader insists on finding."""
    workbook = openpyxl.Workbook()
    monthly = workbook.active
    monthly.title = "Monthly"
    monthly.append(["yyyymm", *GOYAL_WELCH_COLUMNS])
    for key, values in rows.items():
        monthly.append([int(f"{key[0]:04d}{key[1]:02d}"), *values])
    quarterly = workbook.create_sheet("Quarterly")
    quarterly.append(["yyyyq", *GOYAL_WELCH_COLUMNS])
    quarterly.append([19261, *[0.0] * len(GOYAL_WELCH_COLUMNS)])
    annual = workbook.create_sheet("Annual")
    annual.append(["yyyy", *GOYAL_WELCH_COLUMNS])
    annual.append([1926, *[0.0] * len(GOYAL_WELCH_COLUMNS)])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def aqr_xlsx(
    *, sheet: str, columns: Sequence[str], rows: Mapping[tuple[int, int], Sequence[float]]
) -> bytes:
    """An AQR workbook: prose preamble, one header row, month-end date keys, prose tabs."""
    workbook = openpyxl.Workbook()
    data = workbook.active
    data.title = sheet
    data.append(["AQR Capital Management, LLC - synthetic fixture"])
    data.append(["This file contains EXCESS returns and states no cost basis."])
    data.append([None, *columns])
    for key, values in rows.items():
        data.append([month_end(key), *values])
    for name in ("Definitions", "Data Sources", "Disclosures"):
        workbook.create_sheet(name)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.fixture(scope="module")
def generated() -> dict[str, Any]:
    """Every synthetic series, kept as plain arrays so the test can recompute from them."""
    rng = np.random.default_rng(20260816)
    out: dict[str, Any] = {}

    french_keys = month_keys(*COVERAGE["french"])
    market = rng.normal(0.0065, 0.045, len(french_keys)).round(4)
    bill = np.abs(rng.normal(0.0025, 0.001, len(french_keys))).round(4)
    out["french_keys"] = french_keys
    out["market"] = market
    out["french_rows"] = {
        key: (market[i], 0.001, 0.002, bill[i]) for i, key in enumerate(french_keys)
    }

    gw_keys = month_keys(*COVERAGE["goyal_welch"])
    cash = np.abs(rng.normal(0.0028, 0.0012, len(gw_keys))).round(6)
    ltr = (rng.normal(0.0035, 0.028, len(gw_keys)) + cash).round(6)
    corpr = (rng.normal(0.0040, 0.025, len(gw_keys)) + cash).round(6)
    out["gw_keys"] = gw_keys
    out["cash"] = cash
    out["ltr"] = ltr
    out["corpr"] = corpr
    out["gw_rows"] = {
        key: (0.03, cash[i], 0.002, ltr[i], corpr[i]) for i, key in enumerate(gw_keys)
    }

    commodity_keys = month_keys(*COVERAGE["commodity"])
    commodity = rng.normal(0.0025, 0.038, len(commodity_keys)).round(6)
    out["commodity_keys"] = commodity_keys
    out["commodity"] = commodity
    out["commodity_rows"] = {key: (commodity[i],) for i, key in enumerate(commodity_keys)}

    trend_keys = month_keys(*COVERAGE["trend"])
    trend = rng.normal(0.0085, 0.035, len(trend_keys)).round(6)
    out["trend_keys"] = trend_keys
    out["trend"] = trend
    out["trend_rows"] = {key: (trend[i],) for i, key in enumerate(trend_keys)}
    return out


@pytest.fixture(scope="module")
def seeded(
    generated: dict[str, Any], tmp_path_factory: pytest.TempPathFactory
) -> Iterator[dict[str, Any]]:
    """Store the synthetic bytes under the real URLs and return their two hashes each.

    The cache root is exported for the whole module, because the experiment constructs
    its own :class:`RawCache` and must be pointed at the fixture rather than at the
    developer's real 2 GB cache of vendor downloads.
    """
    root = tmp_path_factory.mktemp("cache")
    cache = RawCache(root)
    hashes: dict[str, tuple[str, str]] = {}

    csv_headers = {"content-type": "text/csv", "last-modified": "Mon, 03 Aug 2026 19:17:07 GMT"}
    xlsx_headers = {
        "content-type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "last-modified": "Fri, 26 Jun 2026 15:54:00 GMT",
    }

    dataset = french.get_dataset("french_us_ff3")
    entry = cache.store(dataset.url, french_csv(generated["french_rows"]), headers=csv_headers)
    table = french.parse(cache, entry, dataset=dataset).table("monthly")
    hashes["french_us_ff3"] = (entry.sha256, table.sha256_normalized())

    gw_dataset = goyal_welch.get_dataset("goyal_welch_predictors")
    entry = cache.store(
        gw_dataset.url, goyal_welch_xlsx(generated["gw_rows"]), headers=xlsx_headers
    )
    table = goyal_welch.parse(cache, entry, dataset=gw_dataset).table("monthly")
    hashes["goyal_welch_predictors"] = (entry.sha256, table.sha256_normalized())

    for dataset_id, column, rows_key in (
        ("aqr_commodities_long_run", COMMODITY_COLUMN, "commodity_rows"),
        ("aqr_tsmom_factors", "TSMOM", "trend_rows"),
    ):
        aqr_dataset = aqr.get_dataset(dataset_id)
        payload = aqr_xlsx(
            sheet=aqr_dataset.data_sheet, columns=[column], rows=generated[rows_key]
        )
        entry = cache.store(aqr_dataset.url, payload, headers=xlsx_headers)
        table = aqr.parse(cache, entry, dataset=aqr_dataset).table
        hashes[dataset_id] = (entry.sha256, table.sha256_normalized())

    previous = os.environ.get(CACHE_ENV_VAR)
    os.environ[CACHE_ENV_VAR] = str(root)
    try:
        yield {"root": root, "hashes": hashes}
    finally:
        if previous is None:
            del os.environ[CACHE_ENV_VAR]
        else:
            os.environ[CACHE_ENV_VAR] = previous


def specification_for(
    seeded: dict[str, Any], *, resamples: int = 200, overrides: Mapping[str, str] | None = None
) -> Specification:
    """The committed specification with the synthetic hashes substituted in."""
    raw = yaml.safe_load(default_specification_path().read_text(encoding="utf-8"))
    raw["inference"]["resamples"] = resamples
    for pin in raw["parameters"]["source_pin"]["series"]:
        raw_hash, normalized_hash = seeded["hashes"][pin["dataset_id"]]
        pin["expected_sha256_raw"] = (overrides or {}).get(
            f"{pin['name']}.raw", raw_hash
        )
        pin["expected_sha256_normalized"] = (overrides or {}).get(
            f"{pin['name']}.normalized", normalized_hash
        )
        # The committed manifests describe the real vintages, not these bytes.
        pin["committed_manifest"] = "data-manifests/does-not-exist-for-the-fixture.json"
    return specification_from_mapping(raw, source_path=default_specification_path())


# --------------------------------------------------------------------------- #
# Independent recomputation
# --------------------------------------------------------------------------- #


def expected_panel(generated: dict[str, Any]) -> dict[str, np.ndarray]:
    """Read the panel straight off the generated arrays, by period label."""
    by_label: dict[str, dict[str, float]] = {}

    def put(keys: Sequence[tuple[int, int]], values: np.ndarray, name: str) -> None:
        for key, value in zip(keys, values, strict=True):
            by_label.setdefault(label(key), {})[name] = float(value)

    put(generated["french_keys"], generated["market"], "equity")
    put(generated["gw_keys"], generated["ltr"] - generated["cash"], "treasury")
    put(generated["gw_keys"], generated["corpr"] - generated["cash"], "credit")
    put(generated["gw_keys"], generated["cash"], "cash")
    put(generated["commodity_keys"], generated["commodity"], "commodity")
    put(generated["trend_keys"], generated["trend"], "trend")

    wanted = ("equity", "treasury", "credit", "commodity", "trend", "cash")
    periods = sorted(
        name
        for name, row in by_label.items()
        if all(key in row for key in wanted) and SPEC_WINDOW[0] <= name <= SPEC_WINDOW[1]
    )
    out = {name: np.array([by_label[p][name] for p in periods]) for name in wanted}
    out["periods"] = np.array(periods)
    return out


def expected_portfolio(
    columns: Mapping[str, np.ndarray], weights: Mapping[str, float], charge: float
) -> dict[str, float]:
    excess = sum(weight * columns[name] for name, weight in weights.items()) - charge / 12.0
    total = excess + columns["cash"]
    curve = np.cumprod(1.0 + total)
    peak = np.maximum.accumulate(curve)
    volatility = float(np.std(excess, ddof=1)) * math.sqrt(MONTHS_PER_YEAR)
    return {
        "arithmetic": float(np.mean(excess)) * MONTHS_PER_YEAR,
        "volatility": volatility,
        "sharpe": float(np.mean(excess)) * MONTHS_PER_YEAR / volatility,
        "geometric": float(curve[-1]) ** (MONTHS_PER_YEAR / len(total)) - 1.0,
        "max_drawdown": float(np.min(curve / peak - 1.0)),
    }


# --------------------------------------------------------------------------- #
# The panel
# --------------------------------------------------------------------------- #


def test_the_panel_is_the_intersection_and_neither_boundary_was_chosen(
    seeded: dict[str, Any], generated: dict[str, Any]
) -> None:
    panel = load_panel(specification_for(seeded))
    expected = expected_panel(generated)
    assert panel.periods == tuple(expected["periods"])
    assert panel.periods[0] == "1985-01", "the trend file decides the start"
    assert panel.periods[-1] == "2025-05", "the commodity file decides the end"
    assert panel.months == 485
    assert panel.sleeves == ("equity", "treasury", "credit", "commodity", "trend")


def test_every_sleeve_is_the_series_the_specification_says_it_is(
    seeded: dict[str, Any], generated: dict[str, Any]
) -> None:
    panel = load_panel(specification_for(seeded))
    expected = expected_panel(generated)
    for sleeve in panel.sleeves:
        assert np.allclose(panel.column(sleeve), expected[sleeve], atol=1e-12)
    assert np.allclose(panel.cash, expected["cash"], atol=1e-12)


def test_the_bond_legs_are_differences_taken_inside_one_file(
    seeded: dict[str, Any], generated: dict[str, Any]
) -> None:
    """``ltr - Rfree`` and ``corpr - Rfree`` come from the same row, never two sources."""
    panel = load_panel(specification_for(seeded))
    expected = expected_panel(generated)
    assert np.allclose(
        panel.column("credit") - panel.column("treasury"),
        expected["credit"] - expected["treasury"],
        atol=1e-12,
    )


def test_a_changed_raw_hash_aborts_before_any_statistic_is_computed(
    seeded: dict[str, Any],
) -> None:
    specification = specification_for(seeded, overrides={"trend.raw": "0" * 64})
    with pytest.raises(OverlayStackError, match="NEW VINTAGE"):
        load_panel(specification)


def test_a_changed_normalised_hash_is_a_parser_finding_not_a_hash_to_update(
    seeded: dict[str, Any],
) -> None:
    specification = specification_for(seeded, overrides={"commodity.normalized": "0" * 64})
    with pytest.raises(OverlayStackError, match="parser changed behaviour"):
        load_panel(specification)


def test_the_provenance_carries_both_hashes_for_every_series(seeded: dict[str, Any]) -> None:
    panel = load_panel(specification_for(seeded))
    assert len(panel.provenance) == 6
    for record in panel.provenance:
        assert len(str(record["sha256_raw"])) == 64
        assert len(str(record["sha256_normalized"])) == 64
        assert record["source_url"]


# --------------------------------------------------------------------------- #
# The portfolios
# --------------------------------------------------------------------------- #


@pytest.fixture(scope="module")
def result(seeded: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> ExperimentResult:
    specification = specification_for(seeded)
    context = RunContext(
        run_id="fixture",
        seed=1,
        rng=np.random.default_rng(1),
        artifact_dir=tmp_path_factory.mktemp("artifacts"),
    )
    return run(specification, context)


def _portfolio(result: ExperimentResult, name: str) -> Mapping[str, Any]:
    portfolios = result.diagnostics["portfolios_full_window"]
    assert isinstance(portfolios, Mapping)
    payload = portfolios[name]
    assert isinstance(payload, Mapping)
    return payload


def test_the_unlevered_control_matches_a_plain_numpy_recomputation(
    result: ExperimentResult, generated: dict[str, Any]
) -> None:
    expected = expected_portfolio(expected_panel(generated), {"equity": 1.0}, charge=0.0)
    payload = _portfolio(result, "equity_only")
    assert float(payload["geometric_return"]) == pytest.approx(expected["geometric"])
    assert float(payload["volatility_of_excess_return"]) == pytest.approx(expected["volatility"])
    assert float(payload["sharpe"]) == pytest.approx(expected["sharpe"])
    assert float(payload["max_drawdown"]) == pytest.approx(expected["max_drawdown"])
    assert float(payload["gross_notional"]) == pytest.approx(1.0)
    assert float(payload["annual_cost_charged_on_notional"]) == pytest.approx(0.0)


def test_the_headline_overlay_matches_a_plain_numpy_recomputation(
    result: ExperimentResult, generated: dict[str, Any]
) -> None:
    charge = 0.5 * 0.0145 + 0.0059 * 0.5
    expected = expected_portfolio(
        expected_panel(generated), {"equity": 1.0, "trend": 0.5}, charge=charge
    )
    payload = _portfolio(result, "equity_plus_trend_50")
    assert float(payload["annual_cost_charged_on_notional"]) == pytest.approx(charge)
    assert float(payload["geometric_return"]) == pytest.approx(expected["geometric"])
    assert float(payload["volatility_of_excess_return"]) == pytest.approx(expected["volatility"])
    assert float(payload["sharpe"]) == pytest.approx(expected["sharpe"])
    assert float(payload["gross_notional"]) == pytest.approx(1.5)


def test_the_levered_control_is_charged_the_same_financing_as_the_overlay(
    result: ExperimentResult, generated: dict[str, Any]
) -> None:
    """Exempting it would flatter the overlay by exactly the spread on its notional."""
    charge = 0.0059 * 0.5
    expected = expected_portfolio(expected_panel(generated), {"equity": 1.5}, charge=charge)
    payload = _portfolio(result, "equity_levered_150")
    assert float(payload["annual_cost_charged_on_notional"]) == pytest.approx(charge)
    assert float(payload["geometric_return"]) == pytest.approx(expected["geometric"])
    assert float(payload["sharpe"]) == pytest.approx(expected["sharpe"])
    assert float(payload["sharpe"]) < float(_portfolio(result, "equity_only")["sharpe"])


# --------------------------------------------------------------------------- #
# The comparison
# --------------------------------------------------------------------------- #


def _gaps(result: ExperimentResult, benchmark: str) -> dict[str, Mapping[str, Any]]:
    matched = result.diagnostics["matched_volatility"]
    assert isinstance(matched, Mapping)
    rows = matched[benchmark]
    assert isinstance(rows, Sequence)
    out: dict[str, Mapping[str, Any]] = {}
    for row in rows:
        assert isinstance(row, Mapping)
        out[str(row["portfolio"])] = row
    return out


def test_the_gap_is_the_volatility_times_the_sharpe_difference(
    result: ExperimentResult,
) -> None:
    for benchmark in ("equity_levered_150", "equity_only"):
        for name, row in _gaps(result, benchmark).items():
            expected = float(row["portfolio_volatility"]) * (
                float(row["portfolio_sharpe"]) - float(row["benchmark_sharpe"])
            )
            assert float(row["gap"]) == pytest.approx(expected), name


def test_the_two_benchmarks_are_reported_apart_and_differ_by_the_financing(
    result: ExperimentResult,
) -> None:
    levered = _gaps(result, "equity_levered_150")
    unlevered = _gaps(result, "equity_only")
    assert set(levered) == set(unlevered)
    for name in levered:
        # The levered control's Sharpe is the unlevered one's less the spread it pays,
        # so the gap against it is larger by exactly that much times the portfolio's
        # volatility. The two figures are never added.
        assert float(levered[name]["gap"]) > float(unlevered[name]["gap"])


def test_every_gap_carries_the_smallest_effect_this_sample_could_see(
    result: ExperimentResult,
) -> None:
    for benchmark in ("equity_levered_150", "equity_only"):
        for name, row in _gaps(result, benchmark).items():
            assert float(row["minimum_detectable_effect"]) > 0.0, name
            assert row["resolved"] == (
                abs(float(row["gap"])) >= float(row["minimum_detectable_effect"])
            )


def test_the_haircut_sweep_is_a_straight_line_at_the_sleeve_weight(
    result: ExperimentResult,
) -> None:
    sweeps = result.diagnostics["haircut_sweep"]
    assert isinstance(sweeps, Mapping)
    points = sweeps["equity_levered_150"]
    assert isinstance(points, Sequence)
    gaps = [number(point["gap"]) for point in points if isinstance(point, Mapping)]
    steps = [second - first for first, second in itertools.pairwise(gaps)]
    assert len(steps) == 12
    for step in steps:
        assert step == pytest.approx(-0.5 * 0.01)


def test_the_admission_test_reports_every_sleeve_including_the_one_at_zero_weight(
    result: ExperimentResult,
) -> None:
    rows = result.diagnostics["admission_test"]
    assert isinstance(rows, Sequence)
    sleeves = {str(row["sleeve"]) for row in rows if isinstance(row, Mapping)}
    assert sleeves == {"treasury", "credit", "commodity", "trend"}
    for row in rows:
        assert isinstance(row, Mapping)
        moments = as_mapping(result.diagnostics["moments_full_window"])
        expected = (
            number(row["base_exposure"])
            * number(row["correlation_with_equity"])
            * number(as_mapping(moments["equity"])["volatility"])
        )
        assert number(row["threshold_sharpe"]) == pytest.approx(expected)


# --------------------------------------------------------------------------- #
# What the result may claim
# --------------------------------------------------------------------------- #


def test_the_result_can_never_be_promoted_above_exploratory(result: ExperimentResult) -> None:
    assert result.status in {
        ResultStatus.EXPLORATORY,
        ResultStatus.UNRESOLVED,
        ResultStatus.REJECTED,
    }


def test_the_caveats_state_the_three_things_a_quoted_figure_would_lose(
    result: ExperimentResult,
) -> None:
    joined = " ".join(result.caveats)
    assert "EXPLORATORY AND UNPROMOTABLE" in joined
    assert "gross of the vendor's own trading costs BY OMISSION" in joined
    assert "EXCESS OF CASH, not a collateralised total return" in joined
    assert "never combined" in joined


def test_every_estimate_carries_units_and_either_an_interval_or_a_reason(
    result: ExperimentResult,
) -> None:
    assert result.estimates
    for estimate in result.estimates:
        assert estimate.units.strip()
        assert estimate.interval is not None or estimate.uncertainty_unavailable_reason.strip()


def test_the_growth_figure_is_reported_for_every_portfolio_beside_its_certainty_equivalent(
    result: ExperimentResult,
) -> None:
    names = {estimate.name for estimate in result.estimates}
    for portfolio in ("equity_only", "equity_plus_trend_50", "equity_levered_150"):
        assert f"net_geometric_return[{portfolio}]" in names
        assert f"certainty_equivalent_gamma_3[{portfolio}]" in names


# --------------------------------------------------------------------------- #
# The runner and the ledger
# --------------------------------------------------------------------------- #


def test_the_runner_ledgers_the_attempt_and_writes_hashed_artifacts(
    seeded: dict[str, Any], tmp_path: Path
) -> None:
    ledger = Ledger(tmp_path / "ledger.jsonl")
    outcome = run_experiment(
        specification_for(seeded, resamples=50),
        registry=build_registry(),
        ledger=ledger,
        artifact_root=tmp_path / "artifacts",
        run_id="exp011-fixture",
    )
    assert outcome.status is RunStatus.SUCCEEDED
    assert outcome.artifacts
    events = [entry.event for entry in ledger.read()]
    assert LedgerEvent.STARTED in events
    assert LedgerEvent.SUCCEEDED in events


def test_a_hash_mismatch_is_ledgered_as_a_failure_rather_than_swallowed(
    seeded: dict[str, Any], tmp_path: Path
) -> None:
    ledger = Ledger(tmp_path / "ledger.jsonl")
    with pytest.raises(OverlayStackError):
        run_experiment(
            specification_for(seeded, resamples=10, overrides={"equity.raw": "0" * 64}),
            registry=build_registry(),
            ledger=ledger,
            artifact_root=tmp_path / "artifacts",
            run_id="exp011-abort",
        )
    entries = list(ledger.read())
    assert entries[-1].event is LedgerEvent.FAILED
    assert entries[-1].failure_reason is not None
    assert "NEW VINTAGE" in entries[-1].failure_reason


def test_the_entry_point_resolves_to_this_module(seeded: dict[str, Any]) -> None:
    assert build_registry().resolve(ENTRY_POINT) is run
    assert specification_for(seeded).entry_point == ENTRY_POINT
