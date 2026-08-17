"""Spreadsheet byte readers shared by the long-horizon source adapters.

Three of the sources this repository holds ship a spreadsheet rather than a CSV,
and two spreadsheet formats are involved: the modern zip-based ``.xlsx`` and the
legacy OLE2 ``.xls`` that Robert Shiller has published since 2000. This module
turns raw *bytes* into rows and does nothing else — no unit handling, no period
labelling, no repair. Each provider module keeps its own semantics, exactly as
:mod:`portfolio_edge.data.french`, :mod:`portfolio_edge.data.fred` and
:mod:`portfolio_edge.data.aqr` do.

The container is checked before it is opened. A source that starts returning an
HTML error page, a login wall or a bot-challenge interstitial produces bytes that
are not a spreadsheet at all, and this module raises rather than letting a parser
discover that halfway through a sheet.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from typing import Final

__all__ = [
    "SheetRows",
    "WorkbookParseError",
    "column_unit_string",
    "load_xls_sheets",
    "load_xlsx_sheets",
]

#: Rows of one sheet, in source order, each a tuple of raw cell values.
SheetRows = tuple[tuple[object, ...], ...]


def column_unit_string(columns: Sequence[str], units: Mapping[str, str]) -> str:
    """Render a per-column unit declaration for a genuinely heterogeneous table.

    ``ParsedTable`` carries one ``units`` string, which is right for a file whose
    every number is the same kind of thing — a factor return, a rate — and wrong
    for a predictor workbook that puts an index level, a dollar dividend, a
    book-to-market ratio and an annualised yield in adjacent columns. Declaring
    such a table as ``"mixed"`` says nothing; declaring it column by column says
    exactly what each number is, and because the string is part of
    ``ParsedTable.canonical_bytes`` a change to any column's declared unit changes
    the table digest, which is the behaviour the canonical form exists to provide.

    The rendering is column order, ``name=unit``, pipe-separated, so it is stable
    and diffable. A column absent from ``units`` is rendered ``undeclared``, never
    guessed.
    """
    return "|".join(f"{name}={units.get(name, 'undeclared')}" for name in columns)


_ZIP_MAGIC: Final = b"PK\x03\x04"
_OLE2_MAGIC: Final = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class WorkbookParseError(ValueError):
    """Raised when downloaded bytes are not the spreadsheet they claim to be."""


def _reject_non_container(raw: bytes, *, magic: bytes, label: str, source: str) -> None:
    if raw.startswith(magic):
        return
    head = raw[:200].decode("utf-8", errors="replace").strip()
    raise WorkbookParseError(
        f"{source}: the downloaded bytes are not {label}. First 200 bytes: "
        f"{head!r}. The source has changed format or served an error, a login "
        "wall or a bot-challenge page; record that as a failed acquisition "
        "rather than working around it."
    )


def load_xlsx_sheets(raw: bytes, *, source: str) -> dict[str, SheetRows]:
    """Read every sheet of an ``.xlsx`` workbook into rows of raw cell values.

    Values arrive as openpyxl produces them with ``data_only=True``: numbers as
    ``int`` or ``float``, text as ``str``, dates as ``datetime``, empty cells as
    ``None``. Nothing is coerced here.
    """
    # openpyxl publishes no type stubs and this repository's dependency set is
    # frozen, so the ignore is local rather than a global override that would
    # also silence unreviewed code.
    import openpyxl  # type: ignore[import-untyped]

    _reject_non_container(raw, magic=_ZIP_MAGIC, label="a zip container (.xlsx)", source=source)
    try:
        book = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a wide family of errors here
        raise WorkbookParseError(f"{source}: openpyxl could not open the workbook: {exc}") from exc
    try:
        return {
            str(name): tuple(
                tuple(row) for row in book[name].iter_rows(values_only=True)
            )
            for name in book.sheetnames
        }
    finally:
        book.close()


def load_xls_sheets(raw: bytes, *, source: str) -> dict[str, SheetRows]:
    """Read every sheet of a legacy ``.xls`` (BIFF8 / OLE2) workbook.

    ``xlrd`` 2.x reads ``.xls`` and refuses ``.xlsx`` on purpose, which is why
    both readers exist rather than one. Cells that Excel stored as errors or as
    blanks come back as ``None``; text and numbers come back as ``str`` and
    ``float``. Shiller's workbook writes its own missing marker as the string
    ``"NA"`` and that string is preserved here for the provider module to handle,
    because turning it into ``None`` is a semantic decision and this module makes
    none.
    """
    import xlrd

    _reject_non_container(
        raw, magic=_OLE2_MAGIC, label="an OLE2 compound document (.xls)", source=source
    )
    try:
        book = xlrd.open_workbook(file_contents=raw)
    except Exception as exc:  # xlrd raises several unrelated error types
        raise WorkbookParseError(f"{source}: xlrd could not open the workbook: {exc}") from exc
    sheets: dict[str, SheetRows] = {}
    for name in book.sheet_names():
        sheet = book.sheet_by_name(name)
        sheets[str(name)] = tuple(
            tuple(sheet.row_values(index)) for index in range(sheet.nrows)
        )
    return sheets
