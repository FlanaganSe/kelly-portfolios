"""AQR public data-set reader: direct workbook download and first-party parsing.

Follows :mod:`portfolio_edge.data.french` exactly in contract — fetch the raw
bytes into the cache, parse only from the cache, hash both the raw file and the
derived table, build a manifest, warn loudly, never silently repair — and differs
only where the source differs.

What these workbooks actually look like
---------------------------------------
An AQR data-set download is an ``.xlsx`` with one data sheet and several prose
sheets. The data sheet opens with a variable-length block of disclaimer and
description rows, then a header row, then the observations. The date column holds
Excel serial dates (month-end, not the first of the month), not ISO text.

Three properties drive every decision in this module:

* **The sheet is part of the identity of the data.** AQR changes URLs, workbook
  names, sheet names and revisions, and the same URL served a different workbook
  a year ago. A manifest that pins a sha256 but not a sheet name cannot be
  reproduced, so the sheet actually read is recorded as the first warning on
  every manifest and the parser *refuses* rather than falling back to another
  sheet when the declared one is absent.
* **The row offset is not stable.** AQR edits the header prose. The data block is
  therefore located structurally — the first maximal run of rows whose first cell
  is a date — never by line number, exactly as the French parser does.
* **The methodology is shipped as pictures.** The "Definitions", "Data Sources"
  and "Disclosures" sheets of the time-series-momentum workbook contain nothing
  but their own titles as cell text; their entire content is embedded EMF
  drawings. Anything that reads only cells concludes, wrongly, that the vendor
  documented nothing. :func:`recover_drawing_text` pulls the text records back
  out of the EMF so that the volatility model, the position-sizing rule, the
  instrument universe, the pre-futures splices and the absence of any stated fee
  or transaction-cost basis are all recorded in the manifest rather than lost.
  The recovery is best-effort and says so; it is evidence about the workbook, not
  a substitute for reading it.

Units are **not** inferred from the numbers. The registry declares what this
repository believes a file contains, the parser cross-checks that declaration
against the observed magnitudes, and a contradiction becomes a loud warning. It
never becomes a silent division by 100.
"""

from __future__ import annotations

import datetime as dt
import io
import re
import struct
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Final

from portfolio_edge.data.cache import CacheEntry, RawCache
from portfolio_edge.data.manifest import DatasetManifest, manifest_from_table
from portfolio_edge.data.table import Frequency, ParsedTable

__all__ = [
    "DATASETS",
    "LICENSE_OR_TERMS_URL",
    "PARSER_VERSION",
    "AqrDataset",
    "AqrFile",
    "AqrParseError",
    "AqrSheetMissingError",
    "build_manifests",
    "download",
    "get_dataset",
    "load",
    "parse",
    "recover_drawing_text",
]

#: Bump on any change to parsing behaviour: block detection, sheet selection,
#: period labelling, unit handling, or drawing-text recovery.
PARSER_VERSION: Final = "aqr/1.0.0"

LICENSE_OR_TERMS_URL: Final = "https://www.aqr.com/Insights/Datasets"

_BASE_URL: Final = "https://www.aqr.com/-/media/AQR/Documents/Insights/Data-Sets"

#: Excel's 1900 date system epoch as used by openpyxl once it has converted a
#: serial to a ``datetime``. Kept only so that a naive datetime can be recognised.
_MIN_PLAUSIBLE_YEAR: Final = 1850
_MAX_PLAUSIBLE_YEAR: Final = 2200

#: Day-gap bands used to classify frequency from the observed date column rather
#: than from the file name. A workbook labelled "Monthly" that ships daily rows is
#: exactly the change this catches.
_MONTHLY_GAP_DAYS: Final = (26, 32)
_ANNUAL_GAP_DAYS: Final = (350, 380)
_DAILY_MAX_GAP_DAYS: Final = 5

_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})(?:-(\d{2}))?$")

_AVAILABILITY_MONTHLY: Final = (
    "Posted by the vendor without a published release timestamp or an as-of date "
    "on any row. The HTTP Last-Modified header of the workbook is the only upper "
    "bound this code can observe on when a row became available, and it bounds "
    "the whole file rather than any individual row. The row for month M was "
    "certainly not available during month M. Treat every observation as "
    "unavailable until at least the month after its period."
)

_REVISION_POLICY_RECONSTRUCTED: Final = (
    "Not point-in-time, and the workbook says so itself: 'Data are updated as "
    "they become available. AQR reconstructs the full history each time the "
    "returns are updated.' Every observation back to the first row can therefore "
    "change between downloads, no vintage archive is published, and a sha256 "
    "here identifies the file downloaded rather than what the series looked like "
    "on any earlier date. The series is additionally a vendor-authored historical "
    "reconstruction by a firm that sells the strategy: it is a replication "
    "target, not independent evidence and not an investable product."
)


class AqrParseError(ValueError):
    """Raised when an AQR workbook does not have the expected shape at all."""


class AqrSheetMissingError(AqrParseError):
    """Raised when the declared data sheet is absent from the workbook.

    Deliberately fatal. Silently parsing whichever sheet happens to be first is
    how a workbook revision turns into a changed result that nobody attributes to
    the source.
    """


@dataclass(frozen=True)
class AqrDataset:
    """One public AQR workbook, with its provenance policies.

    Attributes:
        data_sheet: The sheet this repository expects to read. Recorded in every
            manifest, because the sheet is part of the data's identity.
        declared_source_units: What this repository claims the numbers are. It is
            a claim *about* the source, not a statement *by* the source: the
            workbook never declares its units machine-readably. The parser
            cross-checks it against observed magnitudes and warns on a conflict.
        declared_return_basis: Excess or total, and against what. Also a claim.
    """

    dataset_id: str
    filename: str
    data_sheet: str
    description: str
    declared_source_units: str
    declared_units: str
    declared_unit_transform: str
    declared_return_basis: str
    availability_policy: str
    revision_policy: str
    expected_columns: tuple[str, ...] = ()

    @property
    def url(self) -> str:
        return f"{_BASE_URL}/{self.filename}"


DATASETS: Final[dict[str, AqrDataset]] = {
    dataset.dataset_id: dataset
    for dataset in (
        AqrDataset(
            dataset_id="aqr_tsmom_factors",
            filename="Time-Series-Momentum-Factors-Monthly.xlsx",
            data_sheet="TSMOM Factors",
            description=(
                "Monthly excess returns of AQR's diversified time-series-momentum "
                "factor and its four asset-class sub-factors (commodities, equity "
                "indices, fixed income, currencies). An updated and extended "
                "version of the factors in Moskowitz, Ooi and Pedersen (2012), "
                "with construction differences the workbook does not enumerate."
            ),
            declared_source_units="decimal",
            declared_units="decimal",
            declared_unit_transform="identity",
            declared_return_basis=(
                "monthly EXCESS return over cash, per the workbook's own header "
                "row. No fee, transaction-cost, slippage or financing basis is "
                "stated anywhere in the workbook, so the series must be treated as "
                "gross of all of them until the vendor states otherwise."
            ),
            availability_policy=_AVAILABILITY_MONTHLY,
            revision_policy=_REVISION_POLICY_RECONSTRUCTED,
            expected_columns=("TSMOM", "TSMOM^CM", "TSMOM^EQ", "TSMOM^FI", "TSMOM^FX"),
        ),
        AqrDataset(
            dataset_id="aqr_commodities_long_run",
            filename="Commodities-for-the-Long-Run-Index-Level-Data-Monthly.xlsx",
            data_sheet="Commodities for the Long Run",
            description=(
                "Monthly index-level data behind Levine, Ooi, Richardson and "
                "Sasseville, 'Commodities for the Long Run' (FAJ, 2018). The "
                "first column is the EXCESS return over cash of an equal-weight "
                "commodity futures portfolio, decomposed by the workbook into an "
                "excess spot return and an interest-rate-adjusted carry. The "
                "repository's only broad-commodity return series of any kind, "
                "and it starts in 1877."
            ),
            declared_source_units="decimal",
            declared_units="decimal",
            declared_unit_transform="identity",
            declared_return_basis=(
                "monthly EXCESS return over cash, NOT a total return. A fully "
                "collateralised commodity total return is this series plus a "
                "cash return, and the cash series used to re-add is a choice "
                "that must be declared in the experiment specification rather "
                "than made here: the workbook does not state which cash rate it "
                "subtracted. The final two columns are backwardation and "
                "inflation STATE LABELS, not returns, and the last three columns "
                "are excluded from expected_columns for that reason. No fee, "
                "transaction-cost, slippage, roll-cost or financing basis is "
                "stated anywhere in the workbook, so the series is gross of all "
                "of them, which matters more for commodities than for anything "
                "else here: rolling futures is where the costs live."
            ),
            availability_policy=_AVAILABILITY_MONTHLY,
            revision_policy=_REVISION_POLICY_RECONSTRUCTED,
            expected_columns=(
                "Excess return of equal-weight commodities portfolio",
                "Excess spot return of equal-weight commodities portfolio",
                "Interest rate adjusted carry of equal-weight commodities portfolio",
                "Spot return of equal-weight commodities portfolio",
                "Carry of equal-weight commodities portfolio",
                "Excess return of long/short commodities portfolio",
                "Excess spot return of long/short commodities portfolio",
                "Interest rate adjusted carry of long/short commodities portfolio",
                "Aggregate backwardation/contango",
                "State of backwardation/contango",
                "State of inflation",
            ),
        ),
        AqrDataset(
            dataset_id="aqr_credit_risk_premium",
            filename="Credit-Risk-Premium-Preliminary-Paper-Data.xlsx",
            data_sheet="Credit Risk Premium",
            description=(
                "Monthly excess returns behind Asvanunt and Richardson (2015), "
                "'The Credit Risk Premium'. Sourced from Ibbotson's long-term "
                "corporate and long-term government bond total-return series. "
                "A frozen paper vintage: it ends in December 2014 and AQR has "
                "not extended it."
            ),
            declared_source_units="decimal",
            declared_units="decimal",
            declared_unit_transform="identity",
            declared_return_basis=(
                "monthly EXCESS returns, but NOT all against the same benchmark, "
                "which is the trap in this file. GOVT_XS and SP500_XS are excess "
                "of the risk-free rate, so adding a cash return recovers a total "
                "return. CORP_XS is NOT: the paper defines it as the corporate "
                "bond total return less a DURATION-MATCHED government bond "
                "return, estimated by rolling empirical-duration regressions. It "
                "is a duration-hedged credit spread return, and adding cash to it "
                "does not produce a corporate bond total return. Three lines "
                "measured against two different benchmarks must never be summed "
                "or plotted as one ledger."
            ),
            availability_policy=_AVAILABILITY_MONTHLY,
            revision_policy=(
                "Frozen, not updated: this is the preliminary-paper vintage and "
                "the last observation is December 2014. That makes it unusually "
                "reproducible for an AQR file and unusable for anything "
                "concerning the last decade. It is still not point-in-time: the "
                "history was reconstructed once, from the vintages current in "
                "2015, and no earlier vintage is published."
            ),
            expected_columns=("CORP_XS", "GOVT_XS", "SP500_XS"),
        ),
    )
}


def get_dataset(dataset_id: str) -> AqrDataset:
    """Look up a registered dataset, or raise ``KeyError`` naming the choices."""
    try:
        return DATASETS[dataset_id]
    except KeyError:
        raise KeyError(
            f"unknown AQR dataset {dataset_id!r}; known: {sorted(DATASETS)}"
        ) from None


@dataclass(frozen=True)
class AqrFile:
    """Everything parsed out of one AQR workbook.

    Attributes:
        sheet_names: Every sheet in the workbook, in workbook order. Recorded so
            that a later revision which adds, removes or renames a sheet is
            visible without re-downloading.
        data_sheet: The sheet actually parsed.
        preamble: The prose rows above the header, verbatim. They carry the
            vendor's own description of what the numbers are, which is the only
            statement of it in the file.
        table: The one derived table.
        narrative: Text recovered from the workbook's embedded drawings, keyed by
            the media part it came from. Empty when the workbook has none.
        warnings: File-level problems, as opposed to per-table ones.
    """

    sheet_names: tuple[str, ...]
    data_sheet: str
    preamble: str
    table: ParsedTable
    narrative: tuple[tuple[str, str], ...]
    warnings: tuple[str, ...]


def download(
    cache: RawCache,
    dataset: AqrDataset,
    *,
    force: bool = False,
    timeout: float = 60.0,
) -> CacheEntry:
    """Fetch the workbook into ``cache``, reusing cached bytes unless forced.

    No ``User-Agent`` override. Verified on 2026-08-12: the AQR edge serves the
    workbook to the default ``requests`` agent with HTTP 200, so inventing an
    agent string would only add a way for the request to start failing later.
    """
    return cache.fetch(dataset.url, force=force, timeout=timeout)


def parse(cache: RawCache, entry: CacheEntry, *, dataset: AqrDataset) -> AqrFile:
    """Parse a cached AQR workbook.

    Reads exclusively from the cache, so parsing without a stored raw artifact is
    impossible: :meth:`RawCache.read` raises ``RawArtifactMissing``.

    ``dataset`` is required rather than optional, unlike the French reader. There
    is no defensible way to read one of these workbooks without knowing which
    sheet is the data and what the numbers are claimed to be, and guessing either
    is the failure this module exists to prevent.
    """
    raw = cache.read(entry)
    return parse_bytes(raw, dataset=dataset)


def parse_bytes(raw: bytes, *, dataset: AqrDataset) -> AqrFile:
    """Parse workbook bytes. Split out so fixtures can exercise it directly."""
    # No type stubs are published for openpyxl and this repository's dependency
    # set is frozen, so the import is ignored here rather than in a global
    # override that would silence it for code that has not been reviewed.
    import openpyxl  # type: ignore[import-untyped]

    if not raw.startswith(b"PK\x03\x04"):
        raise AqrParseError(
            f"{dataset.dataset_id}: the downloaded bytes are not a zip container, "
            "so they are not an .xlsx workbook. The vendor has changed the format "
            "or served an error page; do not work around this."
        )

    warnings: list[str] = []
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
    except Exception as exc:  # openpyxl raises a wide family of errors here
        raise AqrParseError(
            f"{dataset.dataset_id}: openpyxl could not open the workbook: {exc}"
        ) from exc

    sheet_names = tuple(str(name) for name in workbook.sheetnames)
    if dataset.data_sheet not in sheet_names:
        raise AqrSheetMissingError(
            f"{dataset.dataset_id}: the declared data sheet "
            f"{dataset.data_sheet!r} is not in this workbook. Sheets found: "
            f"{list(sheet_names)}. AQR renames sheets between revisions; freeze a "
            "new dataset entry against the new name rather than letting the "
            "parser pick a different sheet, which would change the numbers "
            "without changing anything a reader sees."
        )

    sheet = workbook[dataset.data_sheet]
    merged = [str(item) for item in sheet.merged_cells.ranges]
    if merged:
        warnings.append(
            f"the data sheet carries {len(merged)} merged cell ranges "
            f"{merged[:6]}. Merged cells were read as their top-left value and "
            "the empty companions as blanks; check the header row if a column "
            "name looks wrong."
        )

    rows = [
        tuple(row)
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True
        )
    ]
    if not rows:
        raise AqrParseError(f"{dataset.dataset_id}: sheet {dataset.data_sheet!r} is empty")

    start, end = _find_data_block(rows, dataset)
    preamble = _preamble_text(rows[: max(start - 1, 0)])
    header_row = rows[start - 1] if start >= 1 else ()

    table = _build_table(rows[start:end], header_row, dataset=dataset, preamble=preamble)

    narrative = recover_drawing_text(raw)
    prose_sheets = _prose_sheet_text_counts(workbook, exclude=dataset.data_sheet)
    if prose_sheets:
        warnings.append(
            "the non-data sheets of this workbook carry almost no readable cell "
            f"text: {dict(prose_sheets)} (sheet name -> substantive text cells). "
            f"{len(narrative)} embedded drawings were found instead, so the "
            "vendor's construction, data sources and disclosures are NOT "
            "machine-readable here. "
            + (
                "Text was recovered from the EMF records on a best-effort basis and "
                "is reproduced in the manifest warnings; it is evidence about the "
                "workbook, not a substitute for reading it."
                if narrative
                else "No drawing text could be recovered, so the vendor's "
                "methodology and disclosures are unavailable from the archived "
                "bytes."
            )
        )

    workbook.close()
    return AqrFile(
        sheet_names=sheet_names,
        data_sheet=dataset.data_sheet,
        preamble=preamble,
        table=table,
        narrative=narrative,
        warnings=tuple(warnings),
    )


def load(
    cache: RawCache, dataset: AqrDataset, *, force: bool = False
) -> tuple[CacheEntry, AqrFile, tuple[DatasetManifest, ...]]:
    """Download if needed, parse, and build manifests, in one call."""
    entry = download(cache, dataset, force=force)
    parsed = parse(cache, entry, dataset=dataset)
    return (entry, parsed, build_manifests(dataset, entry, parsed))


def build_manifests(
    dataset: AqrDataset, entry: CacheEntry, parsed: AqrFile
) -> tuple[DatasetManifest, ...]:
    """Build the manifest for the one table in the workbook.

    The sheet actually read is recorded as a ``SHEET PINNED:`` warning. The
    manifest schema has no sheet field, and adding one would invalidate every
    manifest already committed under schema version 1, so the sheet is pinned
    where a reader cannot miss it instead of where it would be tidy. A manifest
    that does not pin the sheet is not reproducible.
    """
    narrative_notes = tuple(
        f"text recovered from embedded drawing {part} (best-effort, the vendor "
        f"ships this content as a picture rather than as cells): {text}"
        for part, text in parsed.narrative
    )
    return (
        manifest_from_table(
            dataset_id=f"{dataset.dataset_id}_{parsed.table.table_id}",
            entry=entry,
            table=parsed.table,
            parser_version=PARSER_VERSION,
            availability_policy=dataset.availability_policy,
            revision_policy=dataset.revision_policy,
            license_or_terms_url=LICENSE_OR_TERMS_URL,
            extra_warnings=(
                f"SHEET PINNED: {parsed.data_sheet!r}, of sheets "
                f"{list(parsed.sheet_names)}. AQR changes workbook names, sheet "
                "names and revisions, so this pin is part of the data's identity "
                "and a manifest without it is not reproducible.",
                f"return basis claimed by this repository: {dataset.declared_return_basis}",
                (
                    "source file preamble (verbatim, the only place the sheet "
                    f"describes itself): {parsed.preamble.strip()!r}"
                ),
                *parsed.warnings,
                *narrative_notes,
            ),
        ),
    )


# --------------------------------------------------------------------------- #
# Structural parsing
# --------------------------------------------------------------------------- #


def _as_date(value: object) -> dt.date | None:
    """Return the calendar date a cell denotes, or ``None`` if it denotes none.

    Accepts what openpyxl produces for a date-formatted cell and the ISO text a
    revision might ship instead. A bare number is *not* accepted: an unformatted
    Excel serial is indistinguishable from a return, and guessing which one it is
    would silently invent an index.
    """
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        match = _ISO_DATE.match(value.strip())
        if match is None:
            return None
        year, month, day = match.group(1), match.group(2), match.group(3)
        try:
            return dt.date(int(year), int(month), int(day or 1))
        except ValueError:
            return None
    return None


def _find_data_block(
    rows: Sequence[tuple[object, ...]], dataset: AqrDataset
) -> tuple[int, int]:
    """Locate the observations as the longest run of consecutive date-keyed rows.

    Structural, never positional. AQR edits the disclaimer block above the data,
    so any hardcoded offset is wrong by construction; taking the *longest* run
    rather than the first also survives a stray date sitting in the prose.
    """
    runs: list[tuple[int, int]] = []
    index = 0
    total = len(rows)
    while index < total:
        if not rows[index] or _as_date(rows[index][0]) is None:
            index += 1
            continue
        end = index + 1
        while end < total and rows[end] and _as_date(rows[end][0]) is not None:
            end += 1
        runs.append((index, end))
        index = end
    if not runs:
        raise AqrParseError(
            f"{dataset.dataset_id}: sheet {dataset.data_sheet!r} contains no rows "
            "whose first cell is a date. The workbook layout has changed; do not "
            "work around this."
        )
    start, end = max(runs, key=lambda run: run[1] - run[0])
    if start == 0:
        raise AqrParseError(
            f"{dataset.dataset_id}: the data block starts at the first row, so "
            "there is no header row above it and the columns cannot be named."
        )
    return start, end


def _preamble_text(rows: Sequence[tuple[object, ...]]) -> str:
    lines: list[str] = []
    for row in rows:
        cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _classify_frequency(dates: Sequence[dt.date]) -> tuple[Frequency, tuple[str, ...]]:
    """Infer the frequency from the observed date column, and say how."""
    if len(dates) < 2:
        return ("unknown", ("fewer than two observations; frequency not inferable",))
    gaps = sorted((dates[i + 1] - dates[i]).days for i in range(len(dates) - 1))
    median_gap = gaps[len(gaps) // 2]
    if _MONTHLY_GAP_DAYS[0] <= median_gap <= _MONTHLY_GAP_DAYS[1]:
        return ("monthly", ())
    if _ANNUAL_GAP_DAYS[0] <= median_gap <= _ANNUAL_GAP_DAYS[1]:
        return ("annual", ())
    if median_gap <= _DAILY_MAX_GAP_DAYS:
        return ("daily", ())
    return (
        "unknown",
        (
            f"the median gap between observation dates is {median_gap} days, "
            "which matches no expected frequency band; the frequency is recorded "
            "as unknown rather than guessed.",
        ),
    )


def _period_label(date: dt.date, frequency: Frequency) -> str:
    if frequency == "monthly":
        return f"{date.year:04d}-{date.month:02d}"
    if frequency == "annual":
        return f"{date.year:04d}"
    return date.isoformat()


def _column_names(
    header_row: Sequence[object], width: int
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    warnings: list[str] = []
    cells = [None if cell is None else str(cell).strip() for cell in header_row]
    if cells and cells[0]:
        warnings.append(
            f"the header cell above the date column is {cells[0]!r}, not blank; "
            "the vendor has started naming it, which may mean the layout changed."
        )
    names = [cell for cell in cells[1 : width + 1]]
    resolved: list[str] = []
    for position, name in enumerate(names, start=1):
        if name:
            resolved.append(name)
        else:
            resolved.append(f"unnamed_{position}")
            warnings.append(
                f"column {position} of the data block has no header text and was "
                f"named unnamed_{position}. Do not join this table by position."
            )
    while len(resolved) < width:
        position = len(resolved) + 1
        resolved.append(f"unnamed_{position}")
        warnings.append(
            f"the header row declares fewer names than the data block has value "
            f"columns; column {position} was named unnamed_{position}."
        )
    return (tuple(resolved), tuple(warnings))


def _build_table(
    data_rows: Sequence[tuple[object, ...]],
    header_row: Sequence[object],
    *,
    dataset: AqrDataset,
    preamble: str,
) -> ParsedTable:
    warnings: list[str] = []
    dates: list[dt.date] = []
    for row in data_rows:
        date = _as_date(row[0])
        if date is None:  # pragma: no cover - _find_data_block guarantees otherwise
            raise AqrParseError(f"{dataset.dataset_id}: a data row lost its date key")
        if not _MIN_PLAUSIBLE_YEAR <= date.year <= _MAX_PLAUSIBLE_YEAR:
            raise AqrParseError(
                f"{dataset.dataset_id}: observation date {date.isoformat()} is "
                "outside any plausible range, which usually means an unformatted "
                "Excel serial was read as a date."
            )
        dates.append(date)

    width = _value_column_count(data_rows)
    if width == 0:
        raise AqrParseError(
            f"{dataset.dataset_id}: the data block has a date column and no value "
            "columns."
        )
    columns, column_warnings = _column_names(header_row, width)
    warnings.extend(column_warnings)

    frequency, frequency_warnings = _classify_frequency(dates)
    warnings.extend(frequency_warnings)

    values: list[tuple[float | None, ...]] = []
    blank_cells: list[str] = []
    unparsed_cells: list[str] = []
    for date, row in zip(dates, data_rows, strict=True):
        label = _period_label(date, frequency)
        payload = list(row[1 : width + 1]) + [None] * max(0, width - (len(row) - 1))
        parsed_row: list[float | None] = []
        for column, cell in zip(columns, payload[:width], strict=True):
            if cell is None or (isinstance(cell, str) and not cell.strip()):
                blank_cells.append(f"{label}/{column}")
                parsed_row.append(None)
                continue
            if isinstance(cell, bool):
                unparsed_cells.append(f"{label}/{column}={cell!r}")
                parsed_row.append(None)
                continue
            try:
                parsed_row.append(float(cell))  # type: ignore[arg-type]
            except (TypeError, ValueError):
                unparsed_cells.append(f"{label}/{column}={cell!r}")
                parsed_row.append(None)
        values.append(tuple(parsed_row))

    if blank_cells:
        warnings.append(
            f"{len(blank_cells)} cells were empty in the source and became missing "
            f"values, not zeros: {blank_cells[:5]}"
        )
    if unparsed_cells:
        warnings.append(
            f"{len(unparsed_cells)} cells were not numeric and became missing: "
            f"{unparsed_cells[:5]}"
        )

    warnings.extend(
        _unit_findings(values, columns=columns, dataset=dataset, preamble=preamble)
    )

    if dataset.expected_columns and columns != dataset.expected_columns:
        warnings.append(
            f"columns are {list(columns)} but this repository expects "
            f"{list(dataset.expected_columns)}; the vendor has renamed, added or "
            "reordered a series."
        )

    return ParsedTable(
        table_id=frequency if frequency != "unknown" else "observations",
        banner=preamble.splitlines()[0] if preamble.splitlines() else "",
        columns=columns,
        periods=tuple(_period_label(date, frequency) for date in dates),
        values=tuple(values),
        frequency=frequency,
        source_units=dataset.declared_source_units,
        units=dataset.declared_units,
        unit_transform=dataset.declared_unit_transform,
        warnings=tuple(warnings),
    )


def _value_column_count(data_rows: Sequence[tuple[object, ...]]) -> int:
    """Number of columns after the date key that any row actually populates.

    Trailing empty columns are common in these workbooks — the sheet's declared
    dimension runs wider than the data — and counting them would manufacture
    all-missing series.
    """
    last_populated = 0
    for row in data_rows:
        for position in range(len(row) - 1, 0, -1):
            cell = row[position]
            if cell is not None and (not isinstance(cell, str) or cell.strip()):
                last_populated = max(last_populated, position)
                break
    return last_populated


def _unit_findings(
    values: Sequence[tuple[float | None, ...]],
    *,
    columns: Sequence[str],
    dataset: AqrDataset,
    preamble: str,
) -> tuple[str, ...]:
    """Cross-check the registry's units claim against the observed magnitudes.

    This never transforms anything. The workbook declares its units nowhere in
    machine-readable form, so the declaration is a claim this repository makes;
    all this function can do is say loudly when the numbers contradict it. A
    percent file read as decimal, or the reverse, is the error that survives every
    other check because a premium a hundred times too large still has monotonic
    dates and no duplicates.
    """
    numbers = [value for row in values for value in row if value is not None]
    if not numbers:
        return (
            "every value in the table is missing, so the declared units could not "
            "be cross-checked at all.",
        )
    largest = max(abs(value) for value in numbers)
    typical = sorted(abs(value) for value in numbers)[len(numbers) // 2]
    findings = [
        f"units are DECLARED by this repository as {dataset.declared_source_units!r} "
        f"with transform {dataset.declared_unit_transform!r}; the workbook itself "
        "declares no units anywhere in machine-readable form. Observed median "
        f"|value| is {typical:.6g} and the largest is {largest:.6g} across "
        f"{len(columns)} columns."
    ]
    if dataset.declared_units == "decimal" and largest > 1.0:
        findings.append(
            f"CONTRADICTION: the table is declared decimal but holds a value of "
            f"{largest:.6g}. Either the vendor switched to percent or these are "
            "genuine extremes; nothing was transformed, and no statistic computed "
            "from this table is trustworthy until it is resolved."
        )
    if dataset.declared_units == "decimal" and typical > 0.5:
        findings.append(
            f"CONTRADICTION: the table is declared decimal but the median |value| "
            f"is {typical:.6g}, which is the magnitude of a percent series."
        )
    lowered = preamble.lower()
    if "excess return" in lowered:
        findings.append(
            "the sheet preamble states these are EXCESS returns. They are not "
            "total returns and adding a cash rate is required before they can be "
            "compared with a funded portfolio."
        )
    if "percent" in lowered or "%" in preamble:
        findings.append(
            "the sheet preamble mentions a percentage; re-read it before trusting "
            "the declared units."
        )
    return tuple(findings)


#: A prose sheet holding at most this many substantive text cells is reported as
#: carrying its content somewhere other than its cells. Three is generous: the
#: time-series-momentum workbook's Definitions sheet has two boilerplate lines and
#: a title, and its entire methodology in a picture.
_PROSE_SHEET_TEXT_CELLS: Final = 3
_SUBSTANTIVE_TEXT_CHARS: Final = 40


def _prose_sheet_text_counts(workbook: object, *, exclude: str) -> dict[str, int]:
    """Non-data sheets that hold almost no readable cell text, and how little.

    In the time-series-momentum workbook this is Definitions, Data Sources and
    Disclosures: their content is entirely embedded pictures, so a reader that
    looks only at cells concludes the vendor documented nothing.
    """
    names = getattr(workbook, "sheetnames", [])
    counts: dict[str, int] = {}
    for name in names:
        if name == exclude:
            continue
        sheet = workbook[name]  # type: ignore[index]
        substantive = sum(
            1
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None and len(str(cell).strip()) > _SUBSTANTIVE_TEXT_CHARS
        )
        if substantive <= _PROSE_SHEET_TEXT_CELLS:
            counts[str(name)] = substantive
    return counts


# --------------------------------------------------------------------------- #
# Drawing-text recovery
# --------------------------------------------------------------------------- #

#: ``EMR_EXTTEXTOUTW``. The only EMF record type this recovery reads.
_EMR_EXTTEXTOUTW: Final = 84
#: Byte offset of the ``EmrText`` structure inside an ``EMR_EXTTEXTOUTW`` record.
_EMRTEXT_OFFSET: Final = 36
_MAX_RECOVERED_CHARS: Final = 6000


def recover_drawing_text(raw: bytes) -> tuple[tuple[str, str], ...]:
    """Recover text from EMF pictures embedded in an ``.xlsx``.

    Returns ``(media part name, recovered text)`` pairs for every EMF that yields
    any text. Best-effort by construction: it walks the EMF record stream and
    reads only ``EMR_EXTTEXTOUTW`` payloads, so word order follows drawing order
    rather than reading order, and glyph-indexed or vector-outlined text is not
    recoverable at all. It exists because the alternative — recording that the
    vendor documented nothing — is false.
    """
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        return ()
    recovered: list[tuple[str, str]] = []
    with archive:
        for name in sorted(archive.namelist()):
            if not name.lower().endswith(".emf"):
                continue
            text = _emf_text(archive.read(name))
            if text:
                recovered.append((name, text[:_MAX_RECOVERED_CHARS]))
    return tuple(recovered)


def _emf_text(data: bytes) -> str:
    words: list[str] = []
    offset = 0
    total = len(data)
    while offset + 8 <= total:
        record_type, size = struct.unpack_from("<II", data, offset)
        if size < 8 or offset + size > total:
            break
        if record_type == _EMR_EXTTEXTOUTW:
            words.append(_emf_record_text(data[offset : offset + size]))
        offset += size
    joined = " ".join(word for word in words if word)
    return " ".join(joined.split())


def _emf_record_text(record: bytes) -> str:
    if len(record) < _EMRTEXT_OFFSET + 16:
        return ""
    try:
        char_count, string_offset = struct.unpack_from("<II", record, _EMRTEXT_OFFSET + 8)
    except struct.error:  # pragma: no cover - guarded by the length check
        return ""
    stop = string_offset + 2 * char_count
    if string_offset < _EMRTEXT_OFFSET or stop > len(record) or char_count == 0:
        return ""
    text = record[string_offset:stop].decode("utf-16-le", "ignore")
    return "".join(character if character.isprintable() else " " for character in text)
